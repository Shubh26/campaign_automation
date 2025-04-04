import copy
# for reading excel files - http://www.python-excel.org/
import openpyxl
import logging

from utils import file_utils, file_loader
from utils.config_objects import TableInfo, TABLE_LOADING_CONFIG, METADATA, SHEET_NAME, START_ROW, END_ROW, HEADER, \
    ROW_NUMBER, COLUMN_NUMBER


# constants for keys in table_loading_config

# constants for metadata cell


class TableIdentifier(object):
    def __init__(self,**kwargs):
        print("initialized TableIdentifier")

    def identify_tables(self, filepath):
        table_maps = self._identify_tables_get_json_output(filepath)
        return [TableInfo(**table_map) for table_map in table_maps]

    def _identify_tables_get_json_output(self, filepath):
        filetype = file_utils.get_file_type(filepath)
        table_maps = []
        if (filetype=='xlsx' or filetype=="xls"):
            # if filetype is excel then iterate through all sheets
            sheet_names = file_loader.get_sheet_names(filepath)
            for sheet_name in sheet_names:
                table_maps_individual_sheet = self.__identify_tables(file_loader.get_index_row_iterator(filepath, filetype, sheet_name), sheet_name=sheet_name)
                # updating header & start_row if excel file have some merged cells
                table_maps_individual_sheet = [self.__update_header_and_start_row_excel(filepath, table_map) for table_map in table_maps_individual_sheet]
                table_maps.extend(table_maps_individual_sheet)
        else:
            # for csv type
            table_maps = self.__identify_tables(file_loader.get_index_row_iterator(filepath, filetype))
        return  table_maps

    def _is_table_row(self, row, row_index):
        """
        Arguments:
            row: list
                a list with value for each column in that row
            row_index:int
                line no of that row
        return whether it is start of the actual table
        Assumption - If there are more than 1 non empty columns, then it's table row
        """
        non_empty_cols = [v for v in row if len(v.strip()) > 0]
        return len(non_empty_cols) > 1


    def __identify_tables(self, index_row_iterator, sheet_name=None):
        """
        Given an generator which goes throw a file/sheet(sheet in case of excel), this code will identify all tables in that file/sheet
        index_row_iterator:generator or iterator
            This should be generator which returns (line_no, row),
            where row is a list with values from all the columns in that row
            Eg:- (0, ["store_id","Product", "Dollar Sales (non-projected)", "Dollar Sales (non-projected)"...])

        Following are the assumptions this method makes for identifying tables.
        1) Rows with only 1 column (1 non empty column) are metadata rows
        2) Each metadata section is associated with a table. i.e there can't be metadata common to all tables
        3) Metadata section, if present appears before the table
        4) If there are more than 1 non empty columns, then a table starts
            i) all metadata rows accumulated till that point from end of previous table is associated with this table

        """
        table_maps = []
        is_check_new = True # check for a new section, i.e either new metadata or table
        is_new_section = False # whether new section started with this row
        current_start_line_no = -1
        current_end_line_no = -1
        line_no = -1
        is_new_table = False
        is_metadata_row = False # check whether we are currently passing through a metarow
        is_table_started = False # check whether currently we are passing through a table
        table_map = {METADATA: []}
        # list of maps, {'date':2021-07-24}
        metadata = []

        def is_row_empty(row):
            non_empty_cols = [v for v in row if len(v.strip()) > 0]
            return len(non_empty_cols) == 0

        # iterating through data, either csv or xlsx to identify tables
        for line_no, row in index_row_iterator:
            row = list(map(str,row)) # converting cols to strings, some of these could be floats & throw error in subsequent steps
            non_empty_cols = [v for v in row if len(v.strip()) > 0]
            non_emtpy_cols_combined = ','.join(non_empty_cols)
            def __is_new_section():
                # 1st non empty line after an empty line, shows that something is present, could be metadata or actual table
                return not is_row_empty(row) and is_check_new

            if __is_new_section():
                is_new_section = True
                is_check_new = False
                is_table_started = False

            # check if new table start (or metarow check) if
            # 1) either it's a new section start (by section i meant metadata + table)
            # 2) it's still in the metadata section of new section
            if is_new_section or is_metadata_row:
                is_new_table = self._is_table_row(row, line_no) # if there are more than 1 non empty column then it's a table
                is_metadata_row = not is_new_table # If it's a new section & it's not table, it means that these are metadata rows
                is_new_section = False

            def __is_non_empty_metarow():
                return is_metadata_row and not is_row_empty(row)

            if __is_non_empty_metarow():
                # if it's metarow & it have non emtpy values then add values to metadata map
                for (column_number,col) in enumerate(row):
                    col = col.strip()
                    if len(col)>0:
                        # TODO identify the type of metadata & keep the values
                        meta_index = len(metadata)
                        metadata_cell_dict = {"id":f"meta_{meta_index}","value": non_emtpy_cols_combined,
                                              ROW_NUMBER:line_no, COLUMN_NUMBER:column_number}
                        if sheet_name is not None:
                            metadata_cell_dict[SHEET_NAME] = sheet_name
                        metadata.append(metadata_cell_dict)

            if is_new_table:
                # metadata section is over & table is starting
                is_metadata_row = False
                current_start_line_no = line_no
                is_new_table = False
                is_table_started = True

            def __is_end_of_table():
                # empty line, means end of table or it could be a gap in metadata
                # execute this only for the last line
                return is_row_empty(row) and (not is_check_new) and is_table_started

            if __is_end_of_table():
                is_check_new = True
                # -1 because it's zero indexed & if we in an emtpy row now it means that previous row was the last line in the table
                current_end_line_no = line_no - 1
                table_loading_config_dict = {START_ROW: current_start_line_no, END_ROW: current_end_line_no}
                if sheet_name is not None:
                    table_loading_config_dict[SHEET_NAME] = sheet_name
                table_maps.append(
                    {METADATA: metadata,
                     TABLE_LOADING_CONFIG: table_loading_config_dict})
                metadata = []

        def __finished_iteration_without_table_end():
            return not is_check_new
        if __finished_iteration_without_table_end():
            # will enter this block if there was no empty line at the end
            # so we will have to make the entry for the last table
            current_end_line_no = line_no
            table_loading_config_dict = {START_ROW: current_start_line_no, END_ROW: current_end_line_no}
            if sheet_name is not None:
                table_loading_config_dict[SHEET_NAME] = sheet_name
            table_maps.append(
                {METADATA: metadata,
                 TABLE_LOADING_CONFIG: table_loading_config_dict})

        return table_maps

    def __update_header_and_start_row_excel(self, filepath, table_map):
        table_map = copy.deepcopy(table_map)
        start_row = table_map[TABLE_LOADING_CONFIG][START_ROW]
        sheet_name = table_map[TABLE_LOADING_CONFIG][SHEET_NAME]

        # initializing start header & end header to start_row
        start_header = end_header = start_row

        # loading excel file
        wb = openpyxl.load_workbook(filepath, data_only=True)
        work_sheet = wb[sheet_name]

        if len(work_sheet.merged_cells.ranges)>0:
            # work_sheet values have indexes starting with 1, subtracting 1 to make it zero indexed
            merged_cases_with_table_start = [(range_.min_row - 1, range_.max_row - 1) for range_ in
                                             work_sheet.merged_cells.ranges if
                                             ((start_row >= range_.min_row - 1) and (start_row <= range_.max_row - 1))]

            start_header = sorted([e for (s, e) in merged_cases_with_table_start])[0]
            end_header = sorted([e for (s, e) in merged_cases_with_table_start])[-1]

            header_size = end_header - start_header + 1
            if start_header<start_row:
                # if start_header is smaller than start_row it means that there was a row before start row,
                # which is probably part of the same table as there were merged cells including the current start row
                logging.info(f"updating start_row identified from {start_row} to {start_header}")
                table_map[TABLE_LOADING_CONFIG][START_ROW] = start_header

            if header_size > 1:
                # if header_size is 1, then there is no point in keeping header parameter, by default pandas keeps the 1st one as header
                table_map[TABLE_LOADING_CONFIG][HEADER] = list(range(header_size))

        # closing opened excel file
        wb.close()
        del wb
        return table_map