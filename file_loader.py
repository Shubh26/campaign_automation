import pandas as pd
import csv
# for reading excel files - http://www.python-excel.org/
import openpyxl, xlrd
import logging
import copy
import time

from utils import file_utils
from utils.config_objects import MetaDataCell, TABLE_LOADING_CONFIG, END_ROW, HEADER, START_ROW, SHEET_NAME


def get_index_row_iterator(filepath:str, filetype:str="csv", sheet_name=None, **kwargs):
    """
    Given a csv file or xlsx file return a iterator which gives values row by row
    Arguments:
        filepath:str
            filepath to the file
        filetype:str
            supported types csv, xlsx, xls
        sheet_name:str, int
            Applicable for xlsx or xls files
            can pass sheet_name or sheet_index
            Trying to keep a similar behaviour as pandas.read_excel
            https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html
        **kwargs:
            you can pass arguments for openpyxl.load_workbook like read_only=True etc here
    """
    if filetype == "csv":
        with open(filepath) as f:
            delimiter = ","
            QUOTE_CHAR = '"'
            csvreader = csv.reader(f, delimiter=delimiter, quotechar=QUOTE_CHAR)
            for line_no, line in enumerate(csvreader):
                yield (line_no, line)
    elif filetype=="xlsx" or filetype=="xls":
        yield from __get_index_row_iterator_excel(filepath, sheet_name, **kwargs)


def get_index_row_iterator_complete_file(filepath):
    filetype = file_utils.get_file_type(filepath)
    sheet_name = None
    if filetype=="xlsx" or filetype=="xls":
        sheet_names = get_sheet_names(filepath)
        for sheet_name in sheet_names:
            yield from get_index_row_iterator(filepath, filetype, sheet_name)
    else:
        yield from get_index_row_iterator(filepath, filetype, sheet_name)

def __get_sheet_names_slower_using_xlrd(filepath):
    """
    Get sheet_names present in a excel file (xlsx, xls)
    """
    # https://stackoverflow.com/questions/12250024/how-to-obtain-sheet-names-from-xls-files-without-loading-the-whole-file/12250416#12250416
    wb = xlrd.open_workbook(filepath, on_demand=True)
    sheet_names = wb.sheet_names()
    # https://stackoverflow.com/questions/33241837/python-xlrd-book-how-to-close-the-files
    wb.release_resources()
    del wb
    return sheet_names

def get_sheet_names(filepath):
    """
    Get sheet_names present in a excel file (xlsx, xls)
    """
    # https://stackoverflow.com/questions/12250024/how-to-obtain-sheet-names-from-xls-files-without-loading-the-whole-file/12250416#12250416
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_links=False)
    sheet_names = wb.sheetnames
    # # https://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode
    # wb._archive.close()
    wb.close()
    del wb
    return sheet_names

def __get_value_with_merged_cell_lookup(sheet, cell):
    """
    The the value of cell, even if it's a merged cell. For a regular cell will return the cell value
    For a merged cell the value is associated with only the 1st cell in that merged cell
    """
    # https://stackoverflow.com/questions/23562366/how-to-get-value-present-in-a-merged-cell
    applicable_range = [range_ for range_ in sheet.merged_cells.ranges if cell.coordinate in range_]
    cell_value = sheet.cell(applicable_range[0].min_row, applicable_range[0].min_col).value if len(applicable_range)!=0 else cell.value
    return '' if (cell_value is None) else cell_value

def __get_index_row_iterator_excel(filepath:str, sheet_name=None, **kwargs):
    """
    given a excel file & sheet return an iterator which goes through all the data
    Arguments:
        filepath:str
            Filepath to the excel file
        sheet_name:str or int
            either the sheet_name of the sheet_index
        **kwargs:
            you can pass arguments for openpyxl.load_workbook like read_only=True etc here
    """
    if type(sheet_name)==int:
        # getting sheet_name corresponding to the given index
        sheet_names = get_sheet_names(filepath)
        sheet_name = sheet_names[sheet_name]
    start_time = time.time()
    wb = openpyxl.load_workbook(filepath, data_only=True, **kwargs)
    # keeping workbook as read_only will load it faster but it doesn't have attributes to read merged cells. So it can't be used in a lot of cases
    # https://stackoverflow.com/questions/17836986/openpyxl-load-workbook-and-save-take-too-long
    # wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    end_time = time.time()
    logging.debug(f"__get_index_row_iterator_excel worksheet loaded for sheet_name {sheet_name}, time taken {end_time-start_time}")
    work_sheet = wb[sheet_name]
    ignore_merged_cells = False
    try:
        # https://stackoverflow.com/questions/45220078/attributeerror-with-openpyxl
        ranges = work_sheet.merged_cells.ranges
    except Exception as e:
        ignore_merged_cells = True
        logging.error("merged cells not available, hence ignoring them. Pass read_only=False (default) when loading workbook")
    for line_no, row in enumerate(work_sheet.iter_rows()):
        if ignore_merged_cells:
            # merged cells are not available hence ignoring them
            row_only_values = ['' if cell.value is None else cell.value for cell in row]
        else:
            # getting merged cell values as well
            row_only_values = [__get_value_with_merged_cell_lookup(work_sheet, cell) for cell in row]
        yield (line_no, row_only_values)
    wb.close()
    del wb

def _get_config_keys():
    return [TABLE_LOADING_CONFIG]

def get_file_loader_template():
    return {TABLE_LOADING_CONFIG:{}}

def load_dataframe(filepath:str, common_table_config:dict, table_config:dict):
    """
    This function is to load 1 dataframe from a file, given 1 table config
        Arguments:
            filepath:str
                Path to the csv, csv.bz2, csv.gz xlsx or xls file
            common_table_config:dict
                The values under the key - table_loading_config will be used for loading the table
                For a csv file it will accept parameters for pandas.read_csv - https://pandas.pydata.org/docs/reference/api/pandas.read_csv.html
                For a excel file it will accept parameters for pandas.read_excel - https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html
                Apart from these it accepts - start_row, end_row
                 if table goes till end of sheet or file then `end_row` is not mandatory even if start_row is specified

                Commonly used params are -
                    start_row, end_row, sheet_name, header, usecols
                Sample:
                    {
                        "metadata": [],
                        "table_loading_config": {
                            "start_row": 8,
                            "end_row": 100,
                            "sheet_name": 0,
                            "header": [0],
                            "usecols": "B:M"
                        }
                    }
            table_config:dict
                The values under the key - table_loading_config will be used for loading the table
                Refer to the previous param i.e common_table_config description for possible values
    """
    start_time = time.time()
    filetype = file_utils.get_file_type(filepath)
    table_config_to_process = {} if common_table_config is None else common_table_config.get(TABLE_LOADING_CONFIG,{})
    table_config_to_process.update(table_config.get(TABLE_LOADING_CONFIG, {}))
    logging.info(f"table_config_to_process {table_config_to_process}")

    table_config_to_process = copy.deepcopy(table_config_to_process)
    start_row = table_config_to_process.pop(START_ROW, 0)  # removing start_row from dictionary

    # header relative to the start of table
    header = table_config_to_process.get(HEADER, [0])

    if 'skiprows' not in table_config_to_process:
        # TODO skiprows should ideally be not there
        table_config_to_process['skiprows'] = start_row

    if END_ROW in table_config_to_process:
        end_row = table_config_to_process.pop(END_ROW)
        # nrows is the number of non-header rows
        nrows = end_row - start_row - len(header) + 1
        table_config_to_process['nrows'] = nrows

    df = None
    if filetype=='csv' or filetype=='bz2' or filetype=='gz':
        if SHEET_NAME in table_config_to_process:
            del table_config_to_process[SHEET_NAME]
            logging.error(f"please remove sheet_name for file {filepath} as file is in {filetype} format, currently proceeding after removing sheet_name")
        df = pd.read_csv(filepath, **table_config_to_process)

    elif filetype=="xlsx" or filetype=="xls":
        if SHEET_NAME not in table_config_to_process:
            logging.error(f"please provide sheet_name for file {filepath} as file is in {filetype} format, currently proceeding with first sheet")
        df = pd.read_excel(filepath, **table_config_to_process)

    else:
        raise Exception(f"unsupported filetype {filetype} for file {filepath}")
    end_time = time.time()
    logging.debug(f"time taken for load_dataframe {end_time-start_time}")
    return df


def update_metacell_values(filepath:str, metadata_configs:list, **kwargs):
    """
    Update metadata value based on row_number, column_number if "value" is not provided
    Arguments:
        filepath:str
            filepath to the file
        metadata_configs:list
            a list of metadata configs to create a metadata cell
        **kwargs:
            you can pass arguments for openpyxl.load_workbook like read_only=True etc here
            passing read_only=True makes the loading part faster, but merged cell values will be missed
    """
    start_time = time.time()
    filetype = file_utils.get_file_type(filepath)
    metadata_cells = MetaDataCell.create_list_of_metadata(metadata_configs)
    if len(metadata_cells)==0:
        return []
    # Assumption all metadata_cells are from the same sheet
    sheet_names_set = set([metadata_cell.get_sheet_name() for metadata_cell in metadata_cells])
    assert len(sheet_names_set)<=1
    sheet_name = sheet_names_set.pop() if len(sheet_names_set)==1 else None
    index_row_iterator = get_index_row_iterator(filepath, filetype, sheet_name=sheet_name, **kwargs)
    for line_no, row in index_row_iterator:
        is_all_done = all([not metadata_cell.is_empty() for metadata_cell in metadata_cells])
        if(is_all_done):
            # all cells have values now, stopping the iteration
            logging.debug(f"update_metacell_values is_all_done executed {is_all_done}")
            break
        for metadata_cell in metadata_cells:
            # checking if metadata values are present, if not load them based on config
            if metadata_cell.is_empty():
                row_number_cell = metadata_cell.get_row_number()
                if line_no==row_number_cell:
                    column_number_cell = metadata_cell.get_column_number()
                    cell_value = row[column_number_cell] if(column_number_cell<len(row)) else ""
                    metadata_cell.set_value(cell_value)
    end_time = time.time()
    logging.debug(f"time taken for update_metacell_values function {end_time - start_time} for file {filepath}")
    return [metadata_cell.get_dict_format() for metadata_cell in metadata_cells]

class FileLoaderConfigCreator(object):
    def __init__(self,common_config={}, table_configs=[], **kwargs):
        """
        This is used to hold common_config, table_specific configs & update them incrementally
        """
        self.common_config = copy.deepcopy(common_config)
        self.table_configs = copy.deepcopy(table_configs)

    @staticmethod
    def get_commonly_used_table_loader_params():
        """
        Common table_loading_config would look like
        "table_loading_config": {
            "start_row": 8,
            "end_row": 100,
            "sheet_name": 0,
            "header": [0],
            "usecols": "B:M"
        }
        """
        return ["header", "start_row", "end_row", "sheet_name", "usecols"]

    def add_common_loading_config(self, **kwargs):
        """
        Arguments:
            kwargs
                    commonly used loading config params are - "header", "start_row", "end_row", "sheet_name", "usecols"
        """
        table_loading_config = self.common_config.get(TABLE_LOADING_CONFIG, {})
        table_loading_config.update(kwargs)
        self.common_config[TABLE_LOADING_CONFIG] = table_loading_config

    def __assert_valid_table_number(self, table_number):
        assert table_number <= len(
            self.table_configs), f"It's only possible to add transformation for either next table or existing tables, total number of table specific configs so far is {len(self.table_configs)} & table_number is {table_number}"

    def add_loader_config_table_specific(self, table_number, **kwargs):
        """
        Arguments:
            table_number:int
                The table number which we want to use this config to load
            kwargs
                commonly used loading config params are - "header", "start_row", "end_row", "sheet_name", "usecols"

        """
        self.__assert_valid_table_number(table_number)
        table_config = {}
        is_existing_table = table_number < len(self.table_configs)
        if is_existing_table:
            table_config = self.table_configs[table_number]
        table_loading_config = table_config.get(TABLE_LOADING_CONFIG, {})
        table_loading_config.update(kwargs)
        table_config[TABLE_LOADING_CONFIG] = table_loading_config
        if is_existing_table:
            self.table_configs[table_number] = table_config
        else:
            self.table_configs.append(table_config)

    def get_common_config(self):
        return self.common_config

    def get_table_configs(self):
        return self.table_configs
